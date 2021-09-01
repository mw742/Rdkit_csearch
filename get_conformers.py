#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 31 15:48:19 2021

@author: wmm
"""
######################################################################
## For any input InChI, the program will generate conformation
## using Rdkit ETKDG method and then minize each conformer
## and locate the lowest energy conformer using MMFF force filed
######################################################################
 
import sys
import rdkit
from rdkit import Chem
from rdkit.Chem import AllChem
import openpyxl


######################################################################
## Read all the test molecules from an external .txt file,
## and then store in inchi_list
######################################################################
inchi_list=[]
with open("dataset_final.txt", "r") as f:
    for line in f.readlines():
        inchi_list.append(line.strip())
#print(inchi_list)




######################################################################
## Utilize Rdkit to do conformational searching for each molecule
## in the inchi_list, and then minize them in MMFF force filed
## Get the lowest energy conformer and output its energy value,
## 3D structure file
######################################################################

## Set the running time to do conformational generation
## rdkit generate a 3D structure using distance matrix method, 
## and get the final structure through minizing a start structure by ETKDG knowledge-based correct method
## diffent running time means the searching process will start from diffenrent start structures into their nearest minimum
## in this way, hope to scan thorugh the potential energy surface as large as possible
Runtimes=100
limitRMS=2.0


## open an excel file to store all the lowest energy value
data=openpyxl.load_workbook('energy_rdkit.xlsx')
#open the first page of energy_excel file
sheetnames=data.get_sheet_names()
table=data.get_sheet_by_name(sheetnames[0])
table=data.active


## for each molecule in inchi_list, using rdkit distance matrix method to generate local minima, and store in allconf list
for inchi in inchi_list:
    mol=Chem.MolFromInchi(inchi)
    mol=Chem.AddHs(mol)
    allconf=AllChem.EmbedMultipleConfs(mol, numConfs=Runtimes, enforceChirality=True, pruneRmsThresh=limitRMS)
    ## for all the conformers got for one inchi, find the lowest energy conformer by calculating the energy using MMFF force field
    lowest_energy_value=99999
    file_number=inchi_list.index(inchi)
    w=Chem.SDWriter(str(file_number)+'out.sdf')
    #print(len(allconf))
    for confId in allconf:
        #print("confId:", confId)
        ff=AllChem.UFFGetMoleculeForceField(mol, confId=confId)
        ff.Minimize()
        energy_value=ff.CalcEnergy()
        if energy_value < lowest_energy_value:
            lowest_energy_value=energy_value
            lowest_confID=confId
    table.cell(1, file_number+1).value=lowest_energy_value
    ## get the lowest energy 3D structure and remove all the hydrogens
    AllChem.MMFFOptimizeMolecule(mol, confId=lowest_confID) 
    mol_final=Chem.RemoveHs(mol)
    #print("lowest_confID:", lowest_confID)
    #print("lowest_energy_value:", lowest_energy_value)
    w.write(mol_final)
    w.close()


data.save('energy_test.xlsx')



















